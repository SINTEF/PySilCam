import pandas as pd
from PyQt5.QtGui import QPixmap
from PyQt5.QtWidgets import QMainWindow, QApplication, QAction, QStatusBar, QFileDialog, QInputDialog, QMessageBox, \
    QSplashScreen
import pysilcam.postprocess as scpp
import pysilcam.oilgas as scog
from pysilcam.background import correct_im_fast
from pysilcam.fakepymba import silcam_load
from tqdm import tqdm
import numpy as np
import cmocean
import matplotlib.pyplot as plt
import matplotlib
from PyQt5 import QtCore
from PyQt5 import QtWidgets
from matplotlib.backends.backend_qt5agg import FigureCanvasQTAgg as FigureCanvas
import sys
import os
from pysilcam.silcamgui.guicalcs import export_timeseries
from openpyxl import Workbook
from glob import glob
import xarray as xr


class FigFrame(QtWidgets.QFrame):
    '''class for the figure'''
    def __init__(self, parent=None):
        super(FigFrame, self).__init__(parent)
        self.setFrameShape(QtWidgets.QFrame.StyledPanel)
        self.parent = parent
        self.graph_view = PlotView(self)

    def resizeEvent(self, event):
        self.graph_view.setGeometry(self.rect())


class InteractivePlotter(QMainWindow):
    '''main class for this tool'''
    def __init__(self, parent=None):
        super(InteractivePlotter, self).__init__(parent)
        self.showMaximized()
        self.setWindowTitle("SummaryExplorer")
        QApplication.processEvents()
        self.plot_fame = FigFrame(self)

        self.layout = QtWidgets.QVBoxLayout()
        self.layout.addWidget(self.plot_fame)
        self.setLayout(self.layout)
        self.setCentralWidget(self.plot_fame)
        self.showMaximized()
        self.raw_path = ''

        mainMenu = self.menuBar()
        fileMenu = mainMenu.addMenu('File')
        imageMenu = mainMenu.addMenu('Images')

        loadButton = QAction('Load', self)
        loadButton.setStatusTip('Load data')
        loadButton.setShortcut("Ctrl+o")
        loadButton.triggered.connect(self.callLoadData)
        fileMenu.addAction(loadButton)

        self.saveButton = QAction('Save', self)
        self.saveButton.setStatusTip('Save PSD data to xls')
        self.saveButton.setShortcut("Ctrl+s")
        self.saveButton.triggered.connect(self.plot_fame.graph_view.save_data)
        fileMenu.addAction(self.saveButton)
        self.saveButton.setEnabled(False)

        avwinButton = QAction('Average window', self)
        avwinButton.setStatusTip('Change the average window')
        avwinButton.setShortcut("Ctrl+w")
        avwinButton.triggered.connect(self.modify_av_wind)
        fileMenu.addAction(avwinButton)

        self.trimButton = QAction('Trim STATS', self)
        self.trimButton.setStatusTip('Make a STATS.h5 file from the selected region')
        self.trimButton.setShortcut("Ctrl+c")
        self.trimButton.triggered.connect(self.trim_stats)
        fileMenu.addAction(self.trimButton)
        self.trimButton.setEnabled(False)

        exitButton = QAction('Exit', self)
        exitButton.setStatusTip('Close')
        exitButton.triggered.connect(self.close)
        fileMenu.addAction(exitButton)

        self.rawButton = QAction('Raw Image', self)
        self.rawButton.setShortcut("r")
        self.rawButton.triggered.connect(self.find_raw_data)
        imageMenu.addAction(self.rawButton)
        self.rawButton.setEnabled(False)

        self.imcButton = QAction('Corrected Image', self)
        self.imcButton.setShortcut("c")
        self.imcButton.triggered.connect(self.find_imc_data)
        imageMenu.addAction(self.imcButton)
        self.imcButton.setEnabled(False)

        self.toggleButton = QAction('Toggle plot  p', self)
        self.toggleButton.setShortcut("p")
        self.toggleButton.triggered.connect(self.plot_fame.graph_view.toggle_plot)
        self.toggleButton.setEnabled(False)
        mainMenu.addAction(self.toggleButton)

        self.statusBar = QStatusBar()
        self.setStatusBar(self.statusBar)
        self.statusBar.showMessage('Hello. Load a -STATS file to start', 1e12)

    def callLoadData(self):
        self.statusBar.showMessage('Loading data. PLEASE WAIT', 1e12)
        self.plot_fame.graph_view.load_data()
        self.raw_path = ''
        self.raw_files = ''
        self.saveButton.setEnabled(True)
        self.rawButton.setEnabled(True)
        self.imcButton.setEnabled(True)
        self.toggleButton.setEnabled(True)
        self.trimButton.setEnabled(True)
        self.statusBar.clearMessage()
        self.setWindowTitle("SummaryExplorer: " + self.plot_fame.graph_view.stats_filename)

    def keyPressEvent(self, event):
        pressedkey = event.key()
        if (pressedkey == QtCore.Qt.Key_Up) or (pressedkey == QtCore.Qt.Key_W):
            self.plot_fame.graph_view.av_window += pd.Timedelta(seconds=1)
            self.plot_fame.graph_view.update_plot()
            event.accept()
        elif (pressedkey == QtCore.Qt.Key_Right) or (pressedkey == QtCore.Qt.Key_D):
            self.plot_fame.graph_view.mid_time += pd.Timedelta(seconds=1)
            self.plot_fame.graph_view.update_plot()
            event.accept()
        elif (pressedkey == QtCore.Qt.Key_Down) or (pressedkey == QtCore.Qt.Key_S):
            self.plot_fame.graph_view.av_window -= pd.Timedelta(seconds=1)
            self.plot_fame.graph_view.av_window = max(pd.Timedelta(seconds=1), self.plot_fame.graph_view.av_window)
            self.plot_fame.graph_view.update_plot()
            event.accept()
        elif (pressedkey == QtCore.Qt.Key_Left) or (pressedkey == QtCore.Qt.Key_A):
            self.plot_fame.graph_view.mid_time -= pd.Timedelta(seconds=1)
            self.plot_fame.graph_view.av_window = max(pd.Timedelta(seconds=1), self.plot_fame.graph_view.av_window)
            self.plot_fame.graph_view.update_plot()
            event.accept()
        else:
            event.ignore()

    def find_imc_data(self):
        self.extract_filename()
        if len(self.raw_files) == 0:
            return
        bg_window = pd.to_timedelta(5, unit='S')
        start_time = self.plot_fame.graph_view.mid_time - bg_window / 2
        end_time = self.plot_fame.graph_view.mid_time + bg_window / 2
        u = pd.to_datetime(self.plot_fame.graph_view.u)
        midtimeidx = np.argwhere((u >= start_time) & (u < end_time))
        ws = waitsplash()
        self.statusBar.showMessage('Creating background from ' + str(len(midtimeidx)) + ' images', 1e12)
        imbg = np.float64(silcam_load(self.raw_files[midtimeidx[0][0]]))
        for i in range(len(midtimeidx) - 1):
            imbg += np.float64(silcam_load(self.raw_files[midtimeidx[i + 1][0]]))
        imbg /= len(midtimeidx)
        imraw = np.float64(silcam_load(self.filename))
        imc = correct_im_fast(imbg, imraw)
        self.statusBar.showMessage('Background done.', 1e12)
        ws.close()
        self.plot_image(imc)

    def find_raw_data(self):
        self.extract_filename()
        if len(self.raw_files) == 0:
            return
        img = silcam_load(self.filename)
        self.plot_image(img)

    def extract_filename(self):
        if self.raw_path == '':
            self.raw_path = QFileDialog.getExistingDirectory(self,
                                                             caption='Where are the raw data?',
                                                             directory=self.raw_path
                                                             )
            self.raw_files = sorted(glob(os.path.join(self.raw_path,
                                                      '*.silc')))
            if len(self.raw_files) == 0:
                self.raw_files = sorted(glob(os.path.join(self.raw_path,
                                                          '*.bmp')))
            if len(self.raw_files) == 0:
                self.statusBar.showMessage('No data here: ' + self.raw_path, 1e12)
                self.raw_path = ''
                return

        midtimeidx = np.argwhere(self.plot_fame.graph_view.u > self.plot_fame.graph_view.mid_time)[0]
        search_time = self.plot_fame.graph_view.u[midtimeidx].to_pydatetime()[0]
        estimate_filename = os.path.join(self.raw_path,
                                         search_time.strftime('D%Y%m%dT%H%M%S.*.*'))
        filename = glob(estimate_filename)
        if len(filename) == 0:
            print('can''t find this:', estimate_filename)
            return
        self.filename = filename[0]

    def plot_image(self, img):
        cv = FigureCanvas(plt.figure(figsize=(5, 3)))
        cv.setWindowTitle(self.filename)
        plt.imshow(img)
        plt.title(self.filename)
        plt.gca().axis('off')
        cv.show()

    def modify_av_wind(self):
        '''allow the user to modify the averaging period of interest'''
        window_seconds = self.plot_fame.graph_view.av_window.seconds
        input_value, okPressed = QInputDialog.getInt(self, "Get integer", "Average window:", window_seconds, 0, 60 * 60, 1)

        if okPressed:
            self.plot_fame.graph_view.av_window = pd.Timedelta(seconds=input_value)
            if not self.plot_fame.graph_view.stats_filename == '':
                self.plot_fame.graph_view.update_plot()

    def trim_stats(self):
        self.plot_fame.graph_view.save_trimmed_stats()


class PlotView(QtWidgets.QWidget):
    '''class for plotting plots'''
    def __init__(self, parent=None):
        super(PlotView, self).__init__(parent)

        self.fig = plt.figure()
        self.axisconstant = plt.subplot(221)
        self.axispsd = plt.subplot(122)
        self.axistext = plt.subplot(223)
        self.axistext.axis('off')
        self.canvas = FigureCanvas(self.fig)
        self.canvas.setParent(self)

        self.layout = QtWidgets.QVBoxLayout()
        self.layout.addWidget(self.canvas)
        self.layout.setStretchFactor(self.canvas, 1)
        self.setLayout(self.layout)

        self.configfile = ''
        self.stats_filename = ''
        self.stats = []
        self.av_window = pd.Timedelta(seconds=30)
        self.plot_pcolor = 0
        self.datadir = os.getcwd()
        self.canvas.draw()

    def old_data_check(self):
        '''
        checks fo HDF5 STATS and asks user to convert if needed.
        will rename self.stats_filename for future use if needed.

        returns bool which is True on success
        '''
        if not self.stats_filename.endswith('.csv'):
            return True

        h5_file = self.stats_filename.replace('.csv', '.h5')
        if os.path.isfile(h5_file):
            self.stats_filename = h5_file
            return True

        msgBox = QMessageBox()
        msgBox.setText('The STATS data appears to be out-dated csv file.' +
                       '\n\nWe will now load the STATS file ' +
                       'and convert the data to HDF5 files (which might take a while).')
        msgBox.setIcon(QMessageBox.Question)
        msgBox.setWindowTitle('Convert data?')
        msgBox.addButton('OK', QMessageBox.ActionRole)
        msgBox.addButton(QMessageBox.Cancel)
        msgBox.exec_()

        if (msgBox.clickedButton() == QMessageBox.Cancel):
            return False

        ws = waitsplash()
        scpp.statscsv_to_statshdf(self.stats_filename)
        self.stats_filename = h5_file
        ws.close()
        return True

    def load_data(self):
        '''handles loading of data, depending on what is available'''
        self.datadir = os.path.split(self.configfile)[0]

        stats_filename_original = self.stats_filename
        stats_original = self.stats
        self.stats_filename = ''
        self.stats = []
        self.stats_filename = QFileDialog.getOpenFileName(self,
                                                          caption='Load a *-STATS file',
                                                          directory=self.datadir,
                                                          filter='STATS (*-STATS.h5 *-STATS.csv)'
                                                          )[0]
        if self.stats_filename == '':
            self.stats_filename = stats_filename_original
            self.stats = stats_original
            return

        if self.stats_filename.endswith('.h5'):
            timeseriesgas_file = self.stats_filename.replace('-STATS.h5', '-TIMESERIESgas.xlsx')
        else:
            timeseriesgas_file = self.stats_filename.replace('-STATS.csv', '-TIMESERIESgas.xlsx')

        if os.path.isfile(timeseriesgas_file):
            ws = waitsplash()
            self.load_from_timeseries()
            ws.close()
        else:

            if not self.old_data_check():
                return

            msgBox = QMessageBox()
            msgBox.setText('The STATS data appear not to have been exported to TIMSERIES.xlsx' +
                           '\n\nWe will now load the STATS file ' +
                           'and convert the data to TIMSERIES.xls files (which might take a while).' +
                           '\n\nThe next time you load this dataset the xls files will be detected and used for quicker loading.')
            msgBox.setIcon(QMessageBox.Question)
            msgBox.setWindowTitle('Load STATS?')
            load_stats_button = msgBox.addButton('OK',
                                                 QMessageBox.ActionRole)
            msgBox.addButton(QMessageBox.Cancel)
            msgBox.exec_()

            if (msgBox.clickedButton() == load_stats_button):
                if self.configfile == '':
                    self.configfile = QFileDialog.getOpenFileName(self,
                                                                  caption='Load config ini file',
                                                                  directory=self.datadir,
                                                                  filter=(('*.ini'))
                                                                  )[0]
                if self.configfile == '':
                    return
                ws = waitsplash()
                export_timeseries(self.configfile, self.stats_filename)

                self.load_from_timeseries()
                ws.close()
            else:
                return

        self.mid_time = min(self.u) + (max(self.u) - min(self.u)) / 2
        self.setup_figure()

    def load_from_timeseries(self):
        '''uses timeseries xls sheets assuming they are available'''
        filename_base = os.path.splitext(self.stats_filename)[0]
        timeseriesgas_file = filename_base.replace('-STATS', '-TIMESERIESgas.xlsx')
        timeseriesoil_file = filename_base.replace('-STATS', '-TIMESERIESoil.xlsx')

        gas = pd.read_excel(timeseriesgas_file, parse_dates=['Time'])
        oil = pd.read_excel(timeseriesoil_file, parse_dates=['Time'])

        self.dias = np.array(oil.columns[1:53], dtype=float)
        self.vd_oil = oil.iloc[:, 1:53].to_numpy(dtype=float)
        self.vd_gas = gas.iloc[:, 1:53].to_numpy(dtype=float)
        self.vd_total = self.vd_oil + self.vd_gas
        self.u = pd.to_datetime(oil['Time'].values)
        self.d50_gas = gas['D50']
        self.d50_oil = oil['D50']

        self.d50_total = np.zeros_like(self.d50_oil)
        self.cos = np.zeros_like(self.d50_total)
        for i, vd in enumerate(self.vd_total):
            self.d50_total[i] = scpp.d50_from_vd(vd, self.dias)
            self.cos[i] = scog.cos_check(self.dias, self.vd_total[i, :])

    def load_from_stats(self):
        '''loads stats data and converts to timeseries without saving'''
        stats = pd.read_hdf(self.stats_filename, 'ParticleStats/stats')
        stats['timestamp'] = pd.to_datetime(stats['timestamp'])

        u = stats['timestamp'].unique()
        u = pd.to_datetime(u)
        sample_volume = scpp.get_sample_volume(self.settings.PostProcess.pix_size,
                                               path_length=self.settings.PostProcess.path_length)

        dias, bin_lims = scpp.get_size_bins()
        vd_oil = np.zeros((len(u), len(dias)))
        vd_gas = np.zeros_like(vd_oil)
        vd_total = np.zeros_like(vd_oil)
        d50_gas = np.zeros(len(u))
        d50_oil = np.zeros_like(d50_gas)
        d50_total = np.zeros_like(d50_gas)
        self.cos = np.zeros_like(d50_total)
        # @todo make this number of particles per image, and sum according to index later
        nparticles_all = 0
        nparticles_total = 0
        nparticles_oil = 0
        nparticles_gas = 0

        for i, s in enumerate(tqdm(u)):
            substats = stats[stats['timestamp'] == s]
            nparticles_all += len(substats)

            nims = scpp.count_images_in_stats(substats)
            sv = sample_volume * nims

            oil = scog.extract_oil(substats)
            nparticles_oil += len(oil)
            dias, vd_oil_ = scpp.vd_from_stats(oil, self.settings.PostProcess)
            vd_oil_ /= sv
            vd_oil[i, :] = vd_oil_

            gas = scog.extract_gas(substats)
            nparticles_gas += len(gas)
            dias, vd_gas_ = scpp.vd_from_stats(gas, self.settings.PostProcess)
            vd_gas_ /= sv
            vd_gas[i, :] = vd_gas_
            d50_gas[i] = scpp.d50_from_vd(vd_gas_, dias)

            nparticles_total += len(oil) + len(gas)
            vd_total_ = vd_oil_ + vd_gas_
            d50_total[i] = scpp.d50_from_vd(vd_total_, dias)
            vd_total[i, :] = vd_total_

            self.cos[i] = scog.cos_check(dias, vd_total[i, :])

        self.vd_total = vd_total
        self.vd_gas = vd_gas
        self.vd_oil = vd_oil
        self.d50_total = d50_total
        self.d50_oil = d50_oil
        self.d50_gas = d50_gas
        self.u = u.tz_convert(None)
        self.dias = dias
        self.stats = stats

    def setup_figure(self):
        '''sets up the plotting figure'''

        self.axisconstant.clear()
        if self.plot_pcolor == 0:
            self.axisconstant.pcolormesh(self.u, self.dias, np.log(self.vd_total.T), cmap=cmocean.cm.matter, shading='nearest')
            self.axisconstant.plot(self.u, self.d50_total, 'kx', markersize=5, alpha=0.25)
            self.axisconstant.plot(self.u, self.d50_gas, 'bx', markersize=5, alpha=0.25)
            self.axisconstant.set_yscale('log')
            self.axisconstant.set_ylabel('ECD [um]')
            self.axisconstant.set_ylim(10, 12000)
            self.yrange = [1, 12000]
        elif self.plot_pcolor == 1:
            self.axisconstant.plot(self.u, np.sum(self.vd_total, axis=1), 'k.', alpha=0.2)
            self.axisconstant.plot(self.u, np.sum(self.vd_oil, axis=1), '.', color=[0.7, 0.4, 0], alpha=0.2)
            self.axisconstant.plot(self.u, np.sum(self.vd_gas, axis=1), 'b.', alpha=0.2)
            self.yrange = [0, max(np.sum(self.vd_total, axis=1))]
            self.axisconstant.set_ylabel('Volume concentration [uL/L]')
            self.axisconstant.set_yscale('log')
            self.axisconstant.set_ylim(min([min(np.sum(self.vd_total, axis=1)),
                                       min(np.sum(self.vd_oil, axis=1)),
                                       min(np.sum(self.vd_gas, axis=1))]),
                                       max(np.sum(self.vd_total, axis=1)))
        else:
            self.axisconstant.plot(self.u, self.cos, 'k.', alpha=0.2)
            self.yrange = [0, 1]
            self.axisconstant.set_ylabel('Cosine similarity with log-normal')
            self.axisconstant.set_ylim(self.yrange)

        self.start_time = min(self.u)
        self.end_time = max(self.u)
        self.line1 = self.axisconstant.vlines(self.start_time, self.yrange[0], self.yrange[1], 'r')
        self.line2 = self.axisconstant.vlines(self.end_time, self.yrange[0], self.yrange[1], 'r')

        self.fig.canvas.callbacks.connect('button_press_event', self.on_click)

        self.update_plot()

    def on_click(self, event):
        '''if you click the correct place, update the plot based on where you click'''
        if event.inaxes is not None:
            try:
                self.mid_time = pd.to_datetime(matplotlib.dates.num2date(event.xdata)).tz_convert(None)
                self.update_plot()
            except:
                pass
        else:
            pass

    def toggle_plot(self):
        self.plot_pcolor += 1
        if self.plot_pcolor == 3:
            self.plot_pcolor = 0
        self.setup_figure()

    def update_plot(self, save=False):
        '''update the plots and save to excel is save=True'''

        vd_total = xr.DataArray(data=self.vd_total, dims=['time', 'particle_size'], coords=[self.u, self.dias])
        vd_gas = xr.DataArray(data=self.vd_gas, dims=['time', 'particle_size'], coords=[self.u, self.dias])
        vd_oil = xr.DataArray(data=self.vd_oil, dims=['time', 'particle_size'], coords=[self.u, self.dias])

        ds = xr.Dataset()
        ds['vd_total'] = vd_total
        ds['vd_gas'] = vd_gas
        ds['vd_oil'] = vd_oil

        start_time = self.mid_time - self.av_window / 2
        end_time = self.mid_time + self.av_window / 2

        ds_slice = ds.sel(time=slice(start_time, end_time))
        psd_nims = ds_slice.time.size
        if psd_nims < 1:
            self.axispsd.clear()

            string = ''
            string += '\n Num images: {:0.0f}'.format(psd_nims)
            string += '\n Start: ' + str(start_time)
            string += '\n End: ' + str(end_time)
            string += '\n Window [sec.]: {:0.3f}'.format((end_time - start_time).total_seconds())

            self.axistext.clear()
            self.axistext.text(1, 1, string, va='top', ha='right', transform=self.axistext.transAxes)
            self.axistext.axis('off')

            self.line1.remove()
            self.line2.remove()
            self.line1 = self.axisconstant.vlines(start_time, self.yrange[0], self.yrange[1], 'r', linestyle='--')
            self.line2 = self.axisconstant.vlines(end_time, self.yrange[0], self.yrange[1], 'r', linestyle='--')
            self.canvas.draw()
            return

        ds_psd = ds.sel(time=slice(start_time, end_time)).mean(dim='time')

        psd_start = min(ds_slice['time'].values)
        psd_end = max(ds_slice['time'].values)

        psd_vc_total = np.sum(ds_psd['vd_total'].values)
        psd_vc_oil = np.sum(ds_psd['vd_oil'].values)
        psd_vc_gas = np.sum(ds_psd['vd_gas'].values)

        # Use xarray idxmax function to find coordinate label (e.g. size bin) of maximum PSD value
        psd_peak = ds_psd.idxmax()
        psd_peak_total = psd_peak['vd_total'].values
        psd_peak_oil = psd_peak['vd_oil'].values
        psd_peak_gas = psd_peak['vd_gas'].values

        psd_d50_total = scpp.d50_from_vd(ds_psd['vd_total'], self.dias)
        psd_d50_oil = scpp.d50_from_vd(ds_psd['vd_oil'], self.dias)
        psd_d50_gas = scpp.d50_from_vd(ds_psd['vd_gas'], self.dias)

        psd_gor = sum(ds_psd['vd_gas'].values) / (sum(ds_psd['vd_oil'].values) + sum(ds_psd['vd_gas'].values)) * 100

        self.axispsd.clear()
        self.axispsd.plot(self.dias, ds_psd['vd_total'], 'k', linewidth=5, label='Total')
        self.axispsd.plot(self.dias, ds_psd['vd_oil'], color=[0.7, 0.4, 0], label='Oil')
        self.axispsd.plot(self.dias, ds_psd['vd_gas'], 'b', label='Gas')

        self.axispsd.vlines(psd_d50_total, 0, max(ds_psd['vd_total']), 'k', linestyle='--', linewidth=1,
                            label='Total d50: {:0.0f}um'.format(psd_d50_total))
        self.axispsd.vlines(psd_d50_oil, 0, max(ds_psd['vd_oil']), color=[0.7, 0.4, 0], linestyle='--', linewidth=1,
                            label='Oil d50: {:0.0f}um'.format(psd_d50_oil))
        self.axispsd.vlines(psd_d50_gas, 0, max(ds_psd['vd_gas']), 'b', linestyle='--', linewidth=1,
                            label='Gas d50: {:0.0f}um'.format(psd_d50_gas))

        self.axispsd.vlines(psd_peak_total, 0, max(ds_psd['vd_total']), 'k', linestyle=':', linewidth=1,
                            label='Total peak: {:0.0f}um'.format(psd_peak_total))
        self.axispsd.vlines(psd_peak_oil, 0, max(ds_psd['vd_oil']), color=[0.7, 0.4, 0], linestyle=':', linewidth=1,
                            label='Oil peak: {:0.0f}um'.format(psd_peak_oil))
        self.axispsd.vlines(psd_peak_gas, 0, max(ds_psd['vd_gas']), 'b', linestyle=':', linewidth=1,
                            label='Gas peak d50: {:0.0f}um'.format(psd_peak_gas))

        self.axispsd.set_xlabel('ECD [um]')
        self.axispsd.set_ylabel('VD [uL/L]')
        self.axispsd.set_xscale('log')
        self.axispsd.set_xlim(10, 12000)
        self.axispsd.legend(loc='upper left')

        string = ''
        string += 'GOR [%]: {:0.01f}'.format(psd_gor)
        string += '\n\n d50 total [um]: {:0.0f}'.format(psd_d50_total)
        string += '\n peak total [um]: {:0.0f}'.format(psd_peak_total)
        string += '\n d50 oil [um]: {:0.0f}'.format(psd_d50_oil)
        string += '\n peak oil [um]: {:0.0f}'.format(psd_peak_oil)
        string += '\n d50 gas [um]: {:0.0f}'.format(psd_d50_gas)
        string += '\n peak gas [um]: {:0.0f}'.format(psd_peak_gas)
        string += '\n\n VC total [uL/L]: {:0.0f}'.format(psd_vc_total)
        string += '\n VC oil [uL/L]: {:0.0f}'.format(psd_vc_oil)
        string += '\n VC gas [uL/L]: {:0.0f}'.format(psd_vc_gas)
        string += '\n\n Num images: {:0.0f}'.format(psd_nims)
        string += '\n\n Start: ' + str(psd_start)
        string += '\n End: ' + str(psd_end)
        string += '\n Window [sec.] {:0.3f}:'.format(pd.to_timedelta(psd_end - psd_start).total_seconds())
        string += '\n\n mid-time: ' + str(pd.to_datetime(self.mid_time))

        self.axistext.clear()
        self.axistext.text(1, 1, string, va='top', ha='right', transform=self.axistext.transAxes)
        self.axistext.axis('off')

        self.line1.remove()
        self.line2.remove()
        self.line1 = self.axisconstant.vlines(pd.to_datetime(psd_start), self.yrange[0], self.yrange[1], 'r')
        self.line2 = self.axisconstant.vlines(pd.to_datetime(psd_end), self.yrange[0], self.yrange[1], 'r')
        self.canvas.draw()

        if save:
            timestring = pd.to_datetime(psd_start).strftime('D%Y%m%dT%H%M%S')
            outputname = self.stats_filename.replace('-STATS.h5', '-PSD-' + timestring)
            outputname = QFileDialog.getSaveFileName(self,
                                                     "Select file to Save", outputname,
                                                     ".xlsx")
            if outputname[1] == '':
                return
            outputname = outputname[0] + outputname[1]

            wb = Workbook()
            ws = wb.active
            ws['A1'] = 'Start:'
            ws['B1'] = pd.to_datetime(psd_start)
            ws['A2'] = 'Mid:'
            ws['B2'] = self.mid_time
            ws['A3'] = 'End:'
            ws['B3'] = pd.to_datetime(psd_end)

            ws['A5'] = 'Number of images:'
            ws['B5'] = psd_nims

            ws['D5'] = 'd50(microns):'
            ws['E5'] = psd_d50_total
            ws['A6'] = 'Number of particles:'
            ws['B6'] = 'NOT IMPLEMENTED'
            ws['D6'] = 'peak || modal size class (microns):'
            ws['E6'] = psd_peak_total.squeeze()[()]

            ws['D13'] = 'd50(microns):'
            ws['E13'] = psd_d50_oil
            ws['D14'] = 'peak || modal size class (microns):'
            ws['E14'] = psd_peak_oil.squeeze()[()]

            ws['D21'] = 'd50(microns):'
            ws['E21'] = psd_d50_gas
            ws['D22'] = 'peak || modal size class (microns):'
            ws['E22'] = psd_peak_gas.squeeze()[()]

            ws['A8'] = 'Bin mid-sizes (microns):'
            ws['A9'] = 'Vol. Conc. / bin (uL/L):'
            ws['A16'] = 'Vol. Conc. / bin (uL/L):'
            ws['A24'] = 'Vol. Conc. / bin (uL/L):'
            ws['A12'] = 'OIL Info'
            ws['A20'] = 'GAS Info'
            for c in range(len(self.dias)):
                ws.cell(row=8, column=c + 2, value=self.dias[c])
                ws.cell(row=9, column=c + 2, value=ds_psd['vd_total'].values[c])
                ws.cell(row=16, column=c + 2, value=ds_psd['vd_oil'].values[c])
                ws.cell(row=24, column=c + 2, value=ds_psd['vd_gas'].values[c])

            wb.save(outputname)
            print('Saved:', outputname)

    def save_data(self):
        '''call the update_plot function with option to save'''
        self.update_plot(save=True)

    def save_trimmed_stats(self):
        start_time = self.mid_time - self.av_window / 2
        end_time = self.mid_time + self.av_window / 2

        if len(self.stats) == 0:
            reply = QMessageBox.question(self, "STATS file has not been loaded.",
                                         'Would you like to load the STATS file?\n' +
                                         '(It might take some time)\n\n' +
                                         self.stats_filename,
                                         QMessageBox.Ok | QMessageBox.Cancel)
            if reply == QMessageBox.Cancel:
                return

            print('loading ' + self.stats_filename)
            ws = waitsplash()
            self.stats = pd.read_csv(self.stats_filename, parse_dates=['timestamp'])
            ws.close()
            print('loaded ' + self.stats_filename)

        self.trimmed_stats, self.output_filename = scpp.trim_stats(self.stats_filename, start_time, end_time,
                                                                   write_new=False, stats=self.stats)

        if np.isnan(self.trimmed_stats.equivalent_diameter.max()) or len(self.trimmed_stats) == 0:
            QMessageBox.warning(self, "No data in this segment!",
                                'No data was found within the specified time range.',
                                QMessageBox.Ok)
            return

        reply = QMessageBox.question(self, "Save Trimmed STATS file?",
                                     'Would you like to save this file?\n\n' +
                                     self.output_filename,
                                     QMessageBox.Save | QMessageBox.Cancel)
        if reply == QMessageBox.Save:
            print('Saving ' + self.output_filename)
            ws = waitsplash()
            self.trimmed_stats, self.output_filename = scpp.trim_stats(self.stats_filename, start_time, end_time,
                                                                       write_new=True, stats=self.stats)
            ws.close()
            print('New STATS.h5 file written as:', self.output_filename)


class waitsplash():
    def __init__(self):
        path_here = os.path.realpath(__file__)
        imfile = os.path.join(os.path.split(path_here)[0], 'loading.png')
        splash_pix = QPixmap(imfile)
        self.splash = QSplashScreen(splash_pix)
        self.splash.setMask(splash_pix.mask())
        self.splash.show()
        QApplication.processEvents()

    def close(self):
        self.splash.close()


def main():
    app = QApplication(sys.argv)
    window = InteractivePlotter()
    window.show()
    app.exec_()


if __name__ == "__main__":
    main()
