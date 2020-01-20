import pandas as pd
import pysilcam.postprocess as scpp
import pysilcam.oilgas as scog
from pysilcam.config import PySilcamSettings
from pysilcam.background import correct_im_fast
from tqdm import tqdm
import numpy as np
import cmocean
import matplotlib.pyplot as plt
import matplotlib
from PyQt5.QtWidgets import *
from PyQt5 import QtCore
from PyQt5 import QtWidgets
from PyQt5.QtGui import *
from matplotlib.backends.backend_qt4agg import FigureCanvasQTAgg as FigureCanvas
import sys
import os
from pysilcam.silcamgui.guicalcs import export_timeseries
from openpyxl import Workbook
from glob import glob

class FigFrame(QtWidgets.QFrame):
    '''class for the figure'''
    def __init__(self, parent=None):
        super(FigFrame, self).__init__(parent)
        self.setFrameShape(QtWidgets.QFrame.StyledPanel)
        self.parent = parent
        self.graph_view = PlotView(self)

    def resizeEvent(self, event):
        self.graph_view.setGeometry(self.rect())


class InteractivePlotter(QMainWindow):
    '''main class for this tool'''
    def __init__(self, parent=None):
        super(InteractivePlotter, self).__init__(parent)
        self.showMaximized()
        self.setWindowTitle("SummaryExplorer")
        QApplication.processEvents()
        self.plot_fame = FigFrame(self)

        self.layout = QtWidgets.QVBoxLayout()
        self.layout.addWidget(self.plot_fame)
        self.setLayout(self.layout)
        self.setCentralWidget(self.plot_fame)
        self.showMaximized()
        self.raw_path = ''

        mainMenu = self.menuBar()
        fileMenu = mainMenu.addMenu('File')
        imageMenu = mainMenu.addMenu('Images')


        loadButton = QAction('Load', self)
        loadButton.setStatusTip('Load data')
        loadButton.setShortcut("Ctrl+o")
        loadButton.triggered.connect(self.callLoadData)
        fileMenu.addAction(loadButton)

        self.saveButton = QAction('Save', self)
        self.saveButton.setStatusTip('Save PSD data to xls')
        self.saveButton.setShortcut("Ctrl+s")
        self.saveButton.triggered.connect(self.plot_fame.graph_view.save_data)
        fileMenu.addAction(self.saveButton)
        self.saveButton.setEnabled(False)

        avwinButton = QAction('Average window', self)
        avwinButton.setStatusTip('Change the average window')
        avwinButton.setShortcut("Ctrl+w")
        avwinButton.triggered.connect(self.modify_av_wind)
        fileMenu.addAction(avwinButton)

        tswinButton = QAction('Timestep', self)
        tswinButton.setStatusTip('Change the timestep')
        tswinButton.setShortcut("Ctrl+t")
        tswinButton.triggered.connect(self.modify_timestep)
        fileMenu.addAction(tswinButton)

        depminButton = QAction('Depth Range', self)
        depminButton.setStatusTip('Change the min depth for stats')
        depminButton.setShortcut("Ctrl+e")
        depminButton.triggered.connect(self.modify_depth_min)
        fileMenu.addAction(depminButton)

        depmaxButton = QAction('Depth Range', self)
        depmaxButton.setStatusTip('Change the min depth for stats')
        depmaxButton.setShortcut("Ctrl+d")
        depmaxButton.triggered.connect(self.modify_depth_max)
        fileMenu.addAction(depmaxButton)

        maxsizeButton = QAction('Max size', self)
        maxsizeButton.setStatusTip('Change the max size for stats')
        maxsizeButton.setShortcut("Ctrl+h")
        maxsizeButton.triggered.connect(self.modify_maxsize)
        fileMenu.addAction(maxsizeButton)

        minsizeButton = QAction('Min size', self)
        minsizeButton.setStatusTip('Change the min size for stats')
        minsizeButton.setShortcut("Ctrl+g")
        minsizeButton.triggered.connect(self.modify_minsize)
        fileMenu.addAction(minsizeButton)


        ppmaxButton = QAction('profile_plot', self)
        ppmaxButton.setStatusTip('profile_plot')
        ppmaxButton.setShortcut("Ctrl+p")
        ppmaxButton.triggered.connect(self.profile_plot)
        fileMenu.addAction(ppmaxButton)

        exitButton = QAction('Exit', self)
        exitButton.setStatusTip('Close')
        exitButton.triggered.connect(self.close)
        fileMenu.addAction(exitButton)

        self.rawButton = QAction('Raw Image', self)
        self.rawButton.setShortcut("r")
        self.rawButton.triggered.connect(self.find_raw_data)
        imageMenu.addAction(self.rawButton)
        self.rawButton.setEnabled(False)

        self.imcButton = QAction('Corrected Image', self)
        self.imcButton.setShortcut("c")
        self.imcButton.triggered.connect(self.find_imc_data)
        imageMenu.addAction(self.imcButton)
        self.imcButton.setEnabled(False)

        self.toggleButton = QAction('Toggle plot  p', self)
        self.toggleButton.setShortcut("p")
        self.toggleButton.triggered.connect(self.plot_fame.graph_view.toggle_plot)
        self.toggleButton.setEnabled(False)
        mainMenu.addAction(self.toggleButton)

        self.statusBar = QStatusBar()
        self.setStatusBar(self.statusBar)
        self.statusBar.showMessage('Hello. Load a -STATS.csv file to start', 1e12)

    def callLoadData(self):
        self.statusBar.showMessage('Loading data. PLEASE WAIT', 1e12)
        self.plot_fame.graph_view.load_data()
        self.raw_path = ''
        self.raw_files = ''
        self.saveButton.setEnabled(True)
        self.rawButton.setEnabled(True)
        self.imcButton.setEnabled(True)
        self.toggleButton.setEnabled(True)
        self.statusBar.clearMessage()
        self.setWindowTitle("SummaryExplorer: " + self.plot_fame.graph_view.stats_filename)

    def keyPressEvent(self, event):
        pressedkey = event.key()
        if (pressedkey == QtCore.Qt.Key_Up) or (pressedkey == QtCore.Qt.Key_W):
            self.plot_fame.graph_view.av_window += pd.Timedelta(seconds=self.plot_fame.graph_view.timestep)
            self.plot_fame.graph_view.update_plot()
            event.accept()
        elif (pressedkey == QtCore.Qt.Key_Right) or (pressedkey == QtCore.Qt.Key_D):
            self.plot_fame.graph_view.mid_time += pd.Timedelta(seconds=self.plot_fame.graph_view.timestep)
            self.plot_fame.graph_view.update_plot()
            event.accept()
        elif (pressedkey == QtCore.Qt.Key_Down) or (pressedkey == QtCore.Qt.Key_S):
            self.plot_fame.graph_view.av_window -= pd.Timedelta(seconds=self.plot_fame.graph_view.timestep)
            self.plot_fame.graph_view.av_window = max(pd.Timedelta(seconds=self.plot_fame.graph_view.timestep), self.plot_fame.graph_view.av_window)
            self.plot_fame.graph_view.update_plot()
            event.accept()
        elif (pressedkey == QtCore.Qt.Key_Left) or (pressedkey == QtCore.Qt.Key_A):
            self.plot_fame.graph_view.mid_time -= pd.Timedelta(seconds=self.plot_fame.graph_view.timestep)
            self.plot_fame.graph_view.av_window = max(pd.Timedelta(seconds=self.plot_fame.graph_view.timestep), self.plot_fame.graph_view.av_window)
            self.plot_fame.graph_view.update_plot()
            event.accept()
        else:
            event.ignore()

    def find_imc_data(self):
        self.extract_filename()
        if len(self.raw_files) == 0:
            return
        bg_window = pd.to_timedelta(5, unit='S')
        start_time = self.plot_fame.graph_view.mid_time - bg_window / 2
        end_time = self.plot_fame.graph_view.mid_time + bg_window / 2
        u = pd.to_datetime(self.plot_fame.graph_view.u)
        midtimeidx = np.argwhere((u >= start_time) & (u < end_time))
        ws = waitsplash()
        self.statusBar.showMessage('Creating background from ' + str(len(midtimeidx)) + ' images', 1e12)
        imbg = np.float64(np.load(self.raw_files[midtimeidx[0][0]]))
        for i in range(len(midtimeidx)-1):
            imbg += np.float64(np.load(self.raw_files[midtimeidx[i+1][0]]))
        imbg /= len(midtimeidx)
        imraw = np.float64(np.load(self.filename))
        imc = correct_im_fast(imbg, imraw)
        self.statusBar.showMessage('Background done.', 1e12)
        ws.close()
        self.plot_image(imc)

    def find_raw_data(self):
        self.extract_filename()
        if len(self.raw_files) == 0:
            return
        img = np.load(self.filename)
        self.plot_image(img)

    def extract_filename(self):
        if self.raw_path == '':
            self.raw_path = QFileDialog.getExistingDirectory(self,
                                                             caption='Where are the raw data?',
                                                             directory=self.raw_path
                                                             )
            self.raw_files = sorted(glob(os.path.join(self.raw_path,
                                                      '*.silc')))
            if len(self.raw_files) == 0:
                self.raw_files = sorted(glob(os.path.join(self.raw_path,
                                                          '*.bmp')))
            if len(self.raw_files) == 0:
                self.statusBar.showMessage('No data here: ' + self.raw_path, 1e12)
                self.raw_path = ''
                return

        midtimeidx = np.argwhere(self.plot_fame.graph_view.u > self.plot_fame.graph_view.mid_time)[0]
        search_time = self.plot_fame.graph_view.u[midtimeidx].to_pydatetime()[0]
        print('search_time',search_time)
        estimate_filename = os.path.join(self.raw_path,
                                         search_time.strftime('D%Y%m%dT%H%M%S.*.silc'))
        filename = glob(estimate_filename)
        if len(filename)==0:
            print('can''t find this:', estimate_filename)
            return
        self.filename = filename[0]

    def plot_image(self, img):
        cv = FigureCanvas(plt.figure(figsize=(5, 3)))
        cv.setWindowTitle(self.filename)
        plt.imshow(img)
        plt.title(self.filename)
        plt.gca().axis('off')
        cv.show()

    def modify_timestep(self):
        '''allow the user to modify the averaging period of interest'''
        ts = self.plot_fame.graph_view.timestep
        input_value, okPressed = QInputDialog.getInt(self, "Timestep",
                                                    "Timestep (sec):", ts, 0, 300, 1)

        if okPressed:
            self.plot_fame.graph_view.timestep = input_value
            if not self.plot_fame.graph_view.stats_filename == '':
                self.plot_fame.graph_view.update_plot()

    def modify_maxsize(self):
        '''allow the user to modify the averaging period of interest'''
        max_size = self.plot_fame.graph_view.max_size
        input_value, okPressed = QInputDialog.getInt(self, "max size",
                                                     "max size:", max_size, 0, 50000, 1)
        self.plot_fame.graph_view.max_size = input_value
        self.plot_fame.graph_view.max_size = np.max([self.plot_fame.graph_view.min_size,
                                                     self.plot_fame.graph_view.max_size])
        if not self.plot_fame.graph_view.stats_filename == '':
            self.plot_fame.graph_view.update_plot()


    def modify_minsize(self):
        '''allow the user to modify the averaging period of interest'''
        min_size = self.plot_fame.graph_view.min_size
        input_value, okPressed = QInputDialog.getInt(self, "min size",
                                                     "min size:", min_size, 0, 50000, 1)
        self.plot_fame.graph_view.min_size = input_value
        self.plot_fame.graph_view.min_size = np.min([self.plot_fame.graph_view.min_size,
                                                     self.plot_fame.graph_view.max_size])
        if not self.plot_fame.graph_view.stats_filename == '':
            self.plot_fame.graph_view.update_plot()

    def modify_depth_max(self):
        '''allow the user to modify the averaging period of interest'''
        max_depth = self.plot_fame.graph_view.max_depth
        input_value, okPressed = QInputDialog.getInt(self, "Max depth",
                                                    "Max depth:", max_depth, 0, 300, 1)

        if okPressed:
            self.plot_fame.graph_view.max_depth = input_value
            if not self.plot_fame.graph_view.stats_filename == '':
                self.plot_fame.graph_view.update_plot()


    def modify_depth_min(self):
        '''allow the user to modify the averaging period of interest'''
        min_depth = self.plot_fame.graph_view.min_depth
        input_value, okPressed = QInputDialog.getInt(self, "Min depth",
                                                    "Min depth:", min_depth, 0, 300, 1)

        if okPressed:
            self.plot_fame.graph_view.min_depth = input_value
            if not self.plot_fame.graph_view.stats_filename == '':
                self.plot_fame.graph_view.update_plot()


    def modify_av_wind(self):
        '''allow the user to modify the averaging period of interest'''
        window_seconds = self.plot_fame.graph_view.av_window.seconds
        input_value, okPressed = QInputDialog.getInt(self, "Get integer", "Average window:", window_seconds, 0, 60*60, 1)

        if okPressed:
            self.plot_fame.graph_view.av_window = pd.Timedelta(seconds=input_value)
            if not self.plot_fame.graph_view.stats_filename == '':
                self.plot_fame.graph_view.update_plot()

    def profile_plot(self):
        if not self.plot_fame.graph_view.stats_filename == '':
            self.plot_fame.graph_view.update_plot(plot_profile=True)


class PlotView(QtWidgets.QWidget):
    '''class for plotting plots'''
    def __init__(self, parent=None):
        super(PlotView, self).__init__(parent)

        self.fig = plt.figure()
        self.axisconstant = plt.subplot(221)
        self.axispsd = plt.subplot(122)
        self.axistext = plt.subplot(223)
        plt.sca(self.axistext)
        plt.axis('off')
        self.canvas = FigureCanvas(self.fig)
        self.canvas.setParent(self)

        self.layout = QtWidgets.QVBoxLayout()
        self.layout.addWidget(self.canvas)
        self.layout.setStretchFactor(self.canvas, 1)
        self.setLayout(self.layout)

        self.configfile = ''
        self.stats_filename = ''
        self.av_window = pd.Timedelta(seconds=30)
        self.plot_pcolor = 0
        self.datadir = os.getcwd()
        self.canvas.draw()
        self.min_depth = 0
        self.max_depth = 100
        self.min_size = 500
        self.max_size = 50000
        self.timestep = 1


    def load_data(self):
        '''handles loading of data, depending on what is available'''
        self.datadir = os.path.split(self.configfile)[0]

        self.stats_filename = ''
        self.stats_filename = QFileDialog.getOpenFileName(self,
                                                          caption='Load a *-STATS.csv file',
                                                          directory=self.datadir,
                                                          filter=(('*-STATS.csv'))
                                                          )[0]
        if self.stats_filename == '':
            return

        timeseriesgas_file = self.stats_filename.replace('-STATS.csv', '-TIMESERIESgas.xlsx')

        if os.path.isfile(timeseriesgas_file):
            ws = waitsplash()
            self.load_from_timeseries()
            ws.close()
        else:

            msgBox = QMessageBox()
            msgBox.setText('The STATS data appear not to have been exported to TIMSERIES.xlsx' +
                           '\n We can use the STATS file anyway (which might take a while)' +
                           '\n or we can convert the data to TIMSERIES.xls now,'
                           '\n which can be used quickly if you want to load these data another time.')
            msgBox.setIcon(QMessageBox.Question)
            msgBox.setWindowTitle('What to do?')
            load_stats_button = msgBox.addButton('Load stats anyway',
                                                 QMessageBox.ActionRole)
            convert_stats_button = msgBox.addButton('Convert and save timeseries',
                                                    QMessageBox.ActionRole)
            msgBox.addButton(QMessageBox.Cancel)
            msgBox.exec_()
            if self.configfile == '':
                self.configfile = QFileDialog.getOpenFileName(self,
                                                                  caption='Load config ini file',
                                                                  directory=self.datadir,
                                                                  filter=(('*.ini'))
                                                                  )[0]
                if self.configfile == '':
                    return
            if (msgBox.clickedButton() == load_stats_button):
                self.settings = PySilcamSettings(self.configfile)
                self.av_window = pd.Timedelta(seconds=self.settings.PostProcess.window_size)

                ws = waitsplash()
                self.load_from_stats()
                ws.close()

            elif (msgBox.clickedButton() == convert_stats_button):
                ws = waitsplash()
                export_timeseries(self.configfile, self.stats_filename)

                self.load_from_timeseries()
                ws.close()
            else:
                return

        self.mid_time = min(self.u) + (max(self.u) - min(self.u)) / 2
        self.setup_figure()

    def load_from_timeseries(self):
        '''uses timeseries xls sheets assuming they are available'''
        timeseriesgas_file = self.stats_filename.replace('-STATS.csv', '-TIMESERIESmaj.xlsx')
        timeseriesoil_file = self.stats_filename.replace('-STATS.csv', '-TIMESERIESmaj.xlsx')

        gas = pd.read_excel(timeseriesgas_file, parse_dates=['Time'])
        oil = pd.read_excel(timeseriesoil_file, parse_dates=['Time'])

        self.dias = np.array(oil.columns[0:52], dtype=float)
        self.vd_oil = oil.as_matrix(columns=oil.columns[0:52])
        self.vd_gas = gas.as_matrix(columns=gas.columns[0:52])
        self.vd_total = self.vd_oil
        self.u = pd.to_datetime(oil['Time'].values)
        self.d50_gas = gas['D50']
        self.d50_oil = oil['D50']

        lissttime, lisstdepth = get_lisst_depths()
        self.depth = np.interp(np.float64(self.u), np.float64(lissttime), lisstdepth)

        self.d50_total = np.zeros_like(self.d50_oil)
        self.cos = np.zeros_like(self.d50_total)
        for i, vd in enumerate(self.vd_total):
            self.d50_total[i] = scpp.d50_from_vd(vd, self.dias)
            self.cos[i] = scog.cos_check(self.dias, self.vd_total[i, :])


    def load_from_timeseries_og(self):
        '''uses timeseries xls sheets assuming they are available'''
        timeseriesgas_file = self.stats_filename.replace('-STATS.csv', '-TIMESERIESgas.xlsx')
        timeseriesoil_file = self.stats_filename.replace('-STATS.csv', '-TIMESERIESoil.xlsx')

        gas = pd.read_excel(timeseriesgas_file, parse_dates=['Time'])
        oil = pd.read_excel(timeseriesoil_file, parse_dates=['Time'])

        self.dias = np.array(oil.columns[0:52], dtype=float)
        self.vd_oil = oil.as_matrix(columns=oil.columns[0:52])
        self.vd_gas = gas.as_matrix(columns=gas.columns[0:52])
        self.vd_total = self.vd_oil + self.vd_gas
        self.u = pd.to_datetime(oil['Time'].values)
        self.d50_gas = gas['D50']
        self.d50_oil = oil['D50']

        self.d50_total = np.zeros_like(self.d50_oil)
        self.cos = np.zeros_like(self.d50_total)
        for i, vd in enumerate(self.vd_total):
            self.d50_total[i] = scpp.d50_from_vd(vd, self.dias)
            self.cos[i] = scog.cos_check(self.dias, self.vd_total[i,:])


    def load_from_stats(self):
        '''loads stats data and converts to timeseries without saving'''
        stats = pd.read_csv(self.stats_filename, parse_dates=['timestamp'])

        u = stats['timestamp'].unique()
        u = pd.to_datetime(u)
        sample_volume = scpp.get_sample_volume(self.settings.PostProcess.pix_size,
                                               path_length=self.settings.PostProcess.path_length)

        dias, bin_lims = scpp.get_size_bins()
        vd_oil = np.zeros((len(u), len(dias)))
        vd_gas = np.zeros_like(vd_oil)
        vd_total = np.zeros_like(vd_oil)
        d50_gas = np.zeros(len(u))
        d50_oil = np.zeros_like(d50_gas)
        d50_total = np.zeros_like(d50_gas)
        self.cos = np.zeros_like(d50_total)
        # @todo make this number of particles per image, and sum according to index later
        nparticles_all = 0
        nparticles_total = 0
        nparticles_oil = 0
        nparticles_gas = 0

        for i, s in enumerate(tqdm(u)):
            substats = stats[stats['timestamp'] == s]
            nparticles_all += len(substats)

            nims = scpp.count_images_in_stats(substats)
            sv = sample_volume * nims

            oil = scog.extract_oil(substats)
            nparticles_oil += len(oil)
            dias, vd_oil_ = scpp.vd_from_stats(oil, self.settings.PostProcess)
            vd_oil_ /= sv
            vd_oil[i, :] = vd_oil_

            gas = scog.extract_gas(substats)
            nparticles_gas += len(gas)
            dias, vd_gas_ = scpp.vd_from_stats(gas, self.settings.PostProcess)
            vd_gas_ /= sv
            vd_gas[i, :] = vd_gas_
            d50_gas[i] = scpp.d50_from_vd(vd_gas_, dias)

            nparticles_total += len(oil) + len(gas)
            vd_total_ = vd_oil_ + vd_gas_
            d50_total[i] = scpp.d50_from_vd(vd_total_, dias)
            vd_total[i, :] = vd_total_

            self.cos[i] = scog.cos_check(dias, vd_total[i, :])

        self.vd_total = vd_total
        self.vd_gas = vd_gas
        self.vd_oil = vd_oil
        self.d50_total = d50_total
        self.d50_oil = d50_oil
        self.d50_gas = d50_gas
        self.u = u
        self.dias = dias


    def setup_figure(self):
        '''sets up the plotting figure'''
        plt.sca(self.axisconstant)
        plt.cla()
        if self.plot_pcolor==0:
            plt.pcolormesh(self.u, self.dias, np.log(self.vd_total.T), cmap=cmocean.cm.matter)
            # plt.plot(self.u, self.d50_total, 'kx', markersize=5, alpha=0.25)
            # plt.plot(self.u, self.d50_gas, 'bx', markersize=5, alpha=0.25)
            plt.yscale('log')
            plt.ylabel('ECD [um]')
            plt.ylim(10, 12000)
            self.yrange = [1, 12000]
        elif self.plot_pcolor==1:
            plt.plot(self.u, np.sum(self.vd_total,axis=1),'k.', alpha=0.2)
            # plt.plot(self.u, np.sum(self.vd_oil, axis=1), '.', color=[0.7, 0.4, 0], alpha=0.2)
            # plt.plot(self.u, np.sum(self.vd_gas, axis=1), 'b.', alpha=0.2)
            self.yrange = [1, max(np.sum(self.vd_total,axis=1))]
            plt.ylabel('Volume concentration [uL/L]')
            plt.yscale('log')
            plt.ylim(self.yrange)
            # plt.ylim(min([min(np.sum(self.vd_total,axis=1)),
            #               min(np.sum(self.vd_oil,axis=1)),
                          # min(np.sum(self.vd_gas,axis=1))]),
                     # max(np.sum(self.vd_total,axis=1)))
        elif self.plot_pcolor==2:
            plt.plot(self.u, self.cos, 'k.', alpha=0.2)
            self.yrange = [0, 1]
            plt.ylabel('Cosine similarity with log-normal')
            plt.ylim(self.yrange)
        else:
            plt.plot(self.u, self.depth, 'k')
            self.yrange = [np.max(self.depth), 0]
            plt.ylabel('Depth [m]')
            plt.ylim(self.yrange)


        self.start_time = min(self.u)
        self.end_time = max(self.u)
        self.line1 = plt.vlines(self.start_time, self.yrange[0], self.yrange[1], 'r')
        self.line2 = plt.vlines(self.end_time, self.yrange[0], self.yrange[1], 'r')

        self.fig.canvas.callbacks.connect('button_press_event', self.on_click)

        self.update_plot()

    def on_click(self, event):
        '''if you click the correct place, update the plot based on where you click'''
        if event.inaxes is not None:
            try:
                self.mid_time = pd.to_datetime(matplotlib.dates.num2date(event.xdata)).tz_convert(None)
                self.update_plot()
            except:
                pass
        else:
            pass

    def toggle_plot(self):
        self.plot_pcolor += 1
        if self.plot_pcolor==4:
            self.plot_pcolor=0
        self.setup_figure()


    def update_plot(self, save=False, plot_profile=False):
        '''update the plots and save to excel is save=True'''
        start_time = self.mid_time - self.av_window / 2
        end_time = self.mid_time + self.av_window / 2
        u = pd.to_datetime(self.u)
        timeind = np.argwhere((u >= start_time) & (u < end_time) &
                              (self.depth>self.min_depth) & (self.depth<self.max_depth))

        psd_nims = len(timeind)
        if psd_nims < 1:
            plt.sca(self.axispsd)
            plt.cla()

            plt.sca(self.axistext)

            string = ''
            string += '\n Num images: {:0.0f}'.format(psd_nims)
            string += '\n Start: ' + str(start_time)
            string += '\n End: ' + str(end_time)
            string += '\n Window [sec.] {:0.3f}:'.format((end_time - start_time).total_seconds())

            plt.title(string, verticalalignment='top', horizontalalignment='right', loc='right')

            plt.sca(self.axisconstant)
            self.line1.remove()
            self.line2.remove()
            self.line1 = plt.vlines(start_time, self.yrange[0], self.yrange[1], 'r', linestyle='--')
            self.line2 = plt.vlines(end_time, self.yrange[0], self.yrange[1], 'r', linestyle='--')
            self.canvas.draw()
            return

        psd_start = min(u[timeind])
        psd_end = max(u[timeind])

        min_depth = np.min(self.depth[timeind])
        mid_depth = np.mean(self.depth[timeind])
        max_depth = np.max(self.depth[timeind])

        psd_total = np.mean(self.vd_total[timeind, :], axis=0)[0]
        psd_oil = np.mean(self.vd_oil[timeind, :], axis=0)[0]
        psd_gas = np.mean(self.vd_gas[timeind, :], axis=0)[0]

        psd_vc_total = np.sum(psd_total)

        psd_peak_total = self.dias[np.argwhere(psd_total == max(psd_total))][0][0]
        psd_peak_oil = self.dias[np.argwhere(psd_oil == max(psd_oil))][0][0]
        psd_peak_gas = self.dias[np.argwhere(psd_gas == max(psd_gas))][0][0]

        psd_d50_total = scpp.d50_from_vd(psd_total, self.dias)
        psd_d50_oil = scpp.d50_from_vd(psd_oil, self.dias)
        psd_d50_gas = scpp.d50_from_vd(psd_gas, self.dias)

        psd_gor = sum(psd_gas) / (sum(psd_oil) + sum(psd_gas)) * 100

        nd = scpp.vd_to_nd(psd_total, self.dias)
        dind = np.argwhere((self.dias>self.min_size) & (self.dias<self.max_size))
        nc_big = np.sum(nd[dind])
        if self.configfile == '':
            self.configfile = QFileDialog.getOpenFileName(self,
                                                          caption='Load config ini file',
                                                          directory=self.datadir,
                                                          filter=(('*.ini'))
                                                          )[0]
            if self.configfile == '':
                return

        self.settings = PySilcamSettings(self.configfile)

        #sample_volume = scpp.get_sample_volume(self.settings.PostProcess.pix_size,
        #                                       path_length=self.settings.PostProcess.path_length)
        # sample volume scaling must already have been done before conversion to xls
        sample_volume = 1
        nd = scpp.nd_rescale(self.dias, nd, sample_volume=sample_volume)

        junge = scpp.get_j(self.dias, nd)

        ind = np.argwhere(nd > 0)
        nd[ind[0]] = np.nan

        # don't plot zeros
        ind = np.argwhere(nd == 0)
        nd[ind] = np.nan

        plt.sca(self.axispsd)
        plt.cla()
        plt.plot(self.dias, nd, 'k', linewidth=2, label='Total')
        # plt.plot(self.dias, psd_oil, color=[0.7, 0.4, 0], label='Oil')
        # plt.plot(self.dias, psd_gas, 'b', label='Gas')


        plt.vlines(psd_d50_total, 0, max(psd_total), 'k', linestyle='--', linewidth=1, label='Total d50: {:0.0f}um'.format(psd_d50_total))
        # plt.vlines(psd_d50_oil, 0, max(psd_oil), color=[0.7, 0.4, 0], linestyle='--', linewidth=1, label='Oil d50: {:0.0f}um'.format(psd_d50_oil))
        # plt.vlines(psd_d50_gas, 0, max(psd_gas), 'b', linestyle='--', linewidth=1, label='Gas d50: {:0.0f}um'.format(psd_d50_gas))

        plt.vlines(psd_peak_total, 0, max(psd_total), 'k', linestyle=':', linewidth=1, label='Total peak: {:0.0f}um'.format(psd_peak_total))
        # plt.vlines(psd_peak_oil, 0, max(psd_oil), color=[0.7, 0.4, 0], linestyle=':', linewidth=1, label='Oil peak: {:0.0f}um'.format(psd_peak_oil))
        # plt.vlines(psd_peak_gas, 0, max(psd_gas), 'b', linestyle=':', linewidth=1, label='Gas peak d50: {:0.0f}um'.format(psd_peak_gas))

        plt.xlabel('Major Axis Length [um]')
        # plt.ylabel('VD [uL/L]')
        plt.ylabel('Number Concentration [#/L/um]')
        plt.xscale('log')
        plt.yscale('log')
        plt.xlim(10, 12000)
        plt.legend(loc='upper left')

        plt.sca(self.axistext)

        string = ''
        string += 'junge: {:0.1f}'.format(junge)
        string += '\n\n d50 total [um]: {:0.0f}'.format(psd_d50_total)
        string += '\n peak total [um]: {:0.0f}'.format(psd_peak_total)
        string += '\n\n VC total [uL/L]: {:0.1f}'.format(psd_vc_total)
        string += '\n\n NC >' + str(np.round(self.dias[dind[0]])) + '<' +\
                  str(np.round(self.dias[dind[-1]])) + 'um [#/L]: {:0.3f}'.format(nc_big)
        string += '\n min depth [m]: {:0.1f}'.format(min_depth)
        string += '\n mid depth [m]: {:0.1f}'.format(mid_depth)
        string += '\n max depth [m]: {:0.1f}'.format(max_depth)
        string += '\n\n Num images: {:0.0f}'.format(psd_nims)
        string += '\n\n Start: ' + str(pd.to_datetime(psd_start[0]))
        string += '\n End: ' + str(pd.to_datetime(psd_end[0]))
        string += '\n Window [sec.] {:0.3f}:'.format(pd.to_timedelta(psd_end[0]-psd_start[0]).total_seconds())
        string += '\n\n mid-time: ' + str(pd.to_datetime(self.mid_time))

        plt.title(string, verticalalignment='top', horizontalalignment='right', loc='right')

        plt.sca(self.axisconstant)
        self.line1.remove()
        self.line2.remove()
        self.line1 = plt.vlines(pd.to_datetime(psd_start[0]), self.yrange[0], self.yrange[1], 'r')
        self.line2 = plt.vlines(pd.to_datetime(psd_end[0]), self.yrange[0], self.yrange[1], 'r')
        self.canvas.draw()

        if plot_profile:
            dind = np.argwhere((self.dias > self.min_size) & (self.dias < self.max_size)).flatten()
            cv = FigureCanvas(plt.figure(figsize=(5, 3)))
            cv.setWindowTitle('profile')
            plt.plot(np.sum(self.vd_total[timeind, dind], axis=1), self.depth[timeind],'k.', alpha=0.2)
            plt.xlabel('Volume Conc. >' + str(np.round(self.dias[dind[0]])) + '<' +\
                  str(np.round(self.dias[dind[-1]])) + 'um [uL/L]')
            plt.ylabel('Depth [m]')
            plt.ylim(self.max_depth, self.min_depth)
            plt.title('Start: ' + str(pd.to_datetime(psd_start[0])) +
                      '\nEnd: ' + str(pd.to_datetime(psd_end[0])))
            cv.show()

        if save:
            timestring = pd.to_datetime(psd_start[0]).strftime('D%Y%m%dT%H%M%S')
            outputname = self.stats_filename.replace('-STATS.csv','-PSD-' + timestring)
            outputname = QFileDialog.getSaveFileName(self,
                                                   "Select file to Save", outputname,
                                                   ".xlsx")
            if outputname[1]=='':
                return
            outputname = outputname[0] + outputname[1]

            wb = Workbook()
            ws = wb.active
            ws['A1'] = 'Start:'
            ws['B1'] = min(u)
            ws['A2'] = 'Mid:'
            ws['B2'] = self.mid_time
            ws['A3'] = 'End:'
            ws['B3'] = max(u)

            ws['A5'] = 'Number of images:'
            ws['B5'] = psd_nims

            ws['D5'] = 'd50(microns):'
            ws['E5'] = psd_d50_total
            ws['A6'] = 'Number of particles:'
            ws['B6'] = 'NOT IMPLEMENTED'
            ws['D6'] = 'peak || modal size class (microns):'
            ws['E6'] = psd_peak_total

            ws['D13'] = 'd50(microns):'
            ws['E13'] = psd_d50_oil
            ws['D14'] = 'peak || modal size class (microns):'
            ws['E14'] = psd_peak_oil

            ws['D21'] = 'd50(microns):'
            ws['E21'] = psd_d50_gas
            ws['D22'] = 'peak || modal size class (microns):'
            ws['E22'] = psd_peak_gas


            ws['A8'] = 'Bin mid-sizes (microns):'
            ws['A9'] = 'Vol. Conc. / bin (uL/L):'
            ws['A16'] = 'Vol. Conc. / bin (uL/L):'
            ws['A24'] = 'Vol. Conc. / bin (uL/L):'
            ws['A12'] = 'OIL Info'
            ws['A20'] = 'GAS Info'
            # d = ws.cells(row='8')
            for c in range(len(self.dias)):
                ws.cell(row=8, column=c + 2, value=self.dias[c])
                ws.cell(row=9, column=c + 2, value=psd_total[c])
                ws.cell(row=16, column=c + 2, value=psd_oil[c])
                ws.cell(row=24, column=c + 2, value=psd_gas[c])

            wb.save(outputname)
            print('Saved:', outputname)

    def save_data(self):
        '''call the update_plot function with option to save'''
        self.update_plot(save=True)


class waitsplash():
    def __init__(self):
        path_here = os.path.realpath(__file__)
        imfile = os.path.join(os.path.split(path_here)[0], 'loading.png')
        splash_pix = QPixmap(imfile)
        self.splash = QSplashScreen(splash_pix)#, Qt.WindowStaysOnTopHint)
        self.splash.setMask(splash_pix.mask())
        self.splash.show()
        QApplication.processEvents()

    def close(self):
        self.splash.close()


def get_lisst_depths(lisstdir = r"C:\Users\emlynd\Documents\PolarNight\LISST"):

    datafiles = glob(os.path.join(lisstdir, '*.csv'))

    lisst = pd.DataFrame()
    for f in datafiles:
        lisst_ = pd.read_csv(f)
        lisst = lisst.append(lisst_)

    time = pd.to_datetime(lisst['Time'])
    depth = lisst['Depth [m]']
    return time, depth

def main():
    app = QApplication(sys.argv)
    window = InteractivePlotter()
    window.show()
    app.exec_()

if __name__ == "__main__":
    main()